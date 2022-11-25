from openpyxl import load_workbook
from yattag import Doc, indent
  
# Load our Excel File
wb = load_workbook("SAMPLE_XML.xlsx")
# Getting an object of active sheet 1 as invoices
ws = wb.worksheets[0]
# Getting an object of active sheet 1 as invoice item detail
ws1 = wb.worksheets[1]
  
# Returning returns a triplet
doc, tag, text = Doc().tagtext()
  
xml_header = '<?xml version="1.0"?>'
xml_schema = '<NMEXML EximID="4" BranchCode="938493901" ACCOUNTANTCOPYID="">'
  
# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)


with tag('TRANSACTIONS', ('OnError','CONTINUE')):
	#FIRST LOOP ON PARENT which on sheet 1
	for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=33):
		row = [cell.value for cell in row]
		with tag('SALESINVOICE', ('operation','ADD'), ('REQUESTID', '1')):
			with tag("TRANSACTIONID"): text(row[1])
			#Second LOOP ON CHILD which on sheet 2
			for child_row in ws1.iter_rows(min_row=2, max_row=3, min_col=1, max_col=14):
				child_row = [cell.value for cell in child_row]
				if row[2] == child_row[1]:
					with tag("ITEMLINE", ('operation','Add')):
						with tag("KeyID"): text(child_row[2])
						with tag("ITEMNO"): text(child_row[3])
						with tag("QUANTITY"): text(child_row[4])
						with tag("ITEMUNIT"): text(child_row[5])
						with tag("UNITRATIO"): text(child_row[6])
						doc._append("<ITEMRESERVED1/>")
						doc._append("<ITEMRESERVED2/>")
						doc._append("<ITEMRESERVED3/>")
						doc._append("<ITEMRESERVED4/>")
						doc._append("<ITEMRESERVED5/>")
						doc._append("<ITEMRESERVED6/>")
						doc._append("<ITEMRESERVED7/>")
						doc._append("<ITEMRESERVED8/>")
						doc._append("<ITEMRESERVED9/>")
						doc._append("<ITEMRESERVED10/>")
						with tag("ITEMOVDESC"): text(child_row[7])
						with tag("UNITPRICE"): text(child_row[8])
						if child_row[9] == 0:
							with tag("ITEMDISCPC"): text(child_row[9])
						else:
							doc._append("<ITEMDISCPC/>")
						with tag("TAXCODES"): text(child_row[10])
						doc._append("<SOSEQ/>")
						with tag("BRUTOUNITPRICE"): text(child_row[11])
						with tag("WAREHOUSEID"): text(child_row[12])
						with tag("QTYCONTROL"): text(child_row[13])
						doc._append("<DOSEQ/>")
						doc._append("<DOID/>")
			with tag("INVOICENO"): text(row[2])
			with tag("INVOICEDATE"): text(row[3])
			with tag("TAX1ID"): text(row[4])
			with tag("TAX1CODE"): text(row[5])
			doc._append("<TAX2CODE/>")
			with tag("TAX1RATE"): text(row[6])
			with tag("TAX2RATE"): text(row[7])
			with tag("RATE"): text(row[8])
			with tag("INCLUSIVETAX"): text(row[9])
			with tag("CUSTOMERISTAXABLE"): text(row[10])
			with tag("CASHDISCOUNT"): text(row[11])
			if row[12] == 0:
				with tag("CASHDISCPC"): text(child_row[12])
			else:
				doc._append("<CASHDISCPC/>")
			with tag("INVOICEAMOUNT"): text(row[13])
			with tag("FREIGHT"): text(row[14])
			with tag("TERMSID"): text(row[15])
			doc._append("<FOB/>")
			doc._append("<PURCHASEORDERNO/>")
			with tag("WAREHOUSEID"): text(row[16])
			doc._append("<DESCRIPTION/>")
			with tag("SHIPDATE"): text(row[17])
			with tag("DELIVERYORDER"): text(row[18])
			with tag("FISCALRATE"): text(row[19])
			with tag("TAXDATE"): text(row[20])
			with tag("CUSTOMERID"): text(row[21])
			with tag("SALESMANID"): 
				with tag("LASTNAME"): text(row[22])
				with tag("FIRSTNAME"): text(row[23])
			with tag("PRINTED"): text(row[24])
			with tag("SHIPTO1"): text(row[25])
			with tag("SHIPTO2"): text(row[26])
			if row[27] == 0:
				with tag("SHIPTO3"): text(child_row[27])
			else:
				doc._append("<SHIPTO3/>")
			if row[28] == 0:
				with tag("SHIPTO4"): text(child_row[28])
			else:
				doc._append("<SHIPTO4/>")
			if row[29] == 0:
				with tag("SHIPTO5"): text(child_row[29])
			else:
				doc._append("<SHIPTO5/>")
			with tag("ARACCOUNT"): text(row[30])
			with tag("TAXFORMNUMBER"): text(row[31])
			doc._append("<TAXFORMCODE/>")
			with tag("CURRENCYNAME"): text(row[32])
  
result = indent(
	doc.getvalue(),
#	indentation='   ',
#	indent_text=True
)
  
with open("sample.xml", "w") as f:
	f.write(result)